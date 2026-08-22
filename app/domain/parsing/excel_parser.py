"""Pure domain module for parsing Excel/CSV price list files.

No DB, no network — reads bytes and returns structured ImportRowData objects.
All openpyxl work should be wrapped in asyncio.to_thread at the caller level.
"""

from __future__ import annotations

import csv
import io
import re
from collections.abc import Sequence
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook  # type: ignore[import-untyped]

# ---------------------------------------------------------------------------
# Data Models
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ImportRowData:
    """A single parsed row from an Excel/CSV price list."""

    row_no: int
    raw_name: str
    raw_unit: str | None = None
    raw_price: Decimal | None = None
    raw_pack_size: Decimal | None = None
    raw_qty: Decimal | None = None
    extra_columns: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class ParseFileResult:
    """Result of parsing an entire file."""

    rows: list[ImportRowData]
    total_rows: int
    skipped_rows: int
    detected_columns: dict[str, int]  # column_role -> column_index


# ---------------------------------------------------------------------------
# Column Detection
# ---------------------------------------------------------------------------

# Fuzzy column name patterns (uz_latn, uz_cyrl, ru, en)
_COLUMN_PATTERNS: dict[str, list[str]] = {
    "name": [
        r"no[mz]i",
        r"mahsulot",
        r"маҳсулот",
        r"номи",
        r"name",
        r"tovar",
        r"наименование",
        r"название",
        r"товар",
        r"product",
        r"материал",
        r"material",
    ],
    "price": [
        r"narx",
        r"нарх",
        r"price",
        r"цена",
        r"сум",
        r"sum",
        r"so'm",
        r"baho",
        r"стоимость",
        r"cost",
    ],
    "unit": [
        r"birlik",
        r"бирлик",
        r"ўлчов",
        r"unit",
        r"ед\.?\s*изм",
        r"единица",
        r"o'lchov",
        r"ölchov",
    ],
    "pack_size": [
        r"qadoq",
        r"қадоқ",
        r"оғирлик",
        r"fasovka",
        r"фасовка",
        r"pack",
        r"hajm",
        r"объ[её]м",
        r"размер",
        r"og'irlik",
        r"вес",
        r"weight",
    ],
    "qty": [
        r"miqdor",
        r"миқдор",
        r"soni",
        r"сони",
        r"qty",
        r"quantity",
        r"количество",
        r"кол[\-\.]?во",
        r"stock",
        r"zaxira",
    ],
}


def _detect_columns(header_cells: list[str]) -> dict[str, int]:
    """Match header cell values to known column roles via regex patterns.

    Returns a mapping of role -> column index (0-based).
    """
    detected: dict[str, int] = {}
    normalized = [cell.strip().lower() for cell in header_cells]

    for role, patterns in _COLUMN_PATTERNS.items():
        for col_idx, cell_val in enumerate(normalized):
            if not cell_val:
                continue
            for pattern in patterns:
                if re.search(pattern, cell_val):
                    if role not in detected:
                        detected[role] = col_idx
                    break

    return detected


def _parse_decimal(value: object) -> Decimal | None:
    """Safely parse a value into Decimal."""
    if value is None:
        return None
    raw = str(value).strip().replace("\xa0", "").replace(" ", "")
    # Handle comma as decimal separator: "52,000" might be 52000 or 52.000
    # If there's exactly one comma and digits after it are 3, treat as thousands separator
    if "," in raw and "." not in raw:
        parts = raw.split(",")
        if len(parts) == 2 and len(parts[1]) == 3:
            # Thousands separator: "52,000" -> "52000"
            raw = raw.replace(",", "")
        else:
            # Decimal separator: "52,5" -> "52.5"
            raw = raw.replace(",", ".")
    try:
        val = Decimal(raw)
        if val <= 0:
            return None
        return val
    except (InvalidOperation, ValueError):
        return None


# ---------------------------------------------------------------------------
# Excel Parser
# ---------------------------------------------------------------------------


def parse_excel(file_bytes: bytes) -> ParseFileResult:
    """Parse an Excel (.xlsx) file into ImportRowData list.

    Expects the first row to be headers. Auto-detects column mapping.
    Must have at minimum a 'name' column detected to proceed.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return ParseFileResult(rows=[], total_rows=0, skipped_rows=0, detected_columns={})

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return ParseFileResult(rows=[], total_rows=0, skipped_rows=0, detected_columns={})

    # Detect headers from first row
    header_cells = [str(cell) if cell is not None else "" for cell in all_rows[0]]
    detected = _detect_columns(header_cells)

    # If no name column detected, try second row as header
    if "name" not in detected and len(all_rows) > 1:
        header_cells = [str(cell) if cell is not None else "" for cell in all_rows[1]]
        detected = _detect_columns(header_cells)
        data_rows = all_rows[2:]
    else:
        data_rows = all_rows[1:]

    if "name" not in detected:
        # Last resort: assume first column is name, second is price
        detected = {"name": 0}
        if len(header_cells) > 1:
            detected["price"] = 1
        data_rows = all_rows  # No header detected, treat all as data

    return _parse_data_rows(data_rows, detected)


# ---------------------------------------------------------------------------
# CSV Parser
# ---------------------------------------------------------------------------


def parse_csv(file_bytes: bytes, encoding: str = "utf-8") -> ParseFileResult:
    """Parse a CSV file into ImportRowData list.

    Tries specified encoding first, falls back to cp1251 (common for Russian).
    """
    text: str | None = None
    for enc in [encoding, "cp1251", "utf-8-sig", "latin-1"]:
        try:
            text = file_bytes.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue

    if text is None:
        return ParseFileResult(rows=[], total_rows=0, skipped_rows=0, detected_columns={})

    # Detect delimiter
    sniffer = csv.Sniffer()
    try:
        dialect: type[csv.Dialect] | csv.Dialect = sniffer.sniff(text[:2048])
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    all_rows = list(reader)

    if not all_rows:
        return ParseFileResult(rows=[], total_rows=0, skipped_rows=0, detected_columns={})

    header_cells = all_rows[0]
    detected = _detect_columns(header_cells)

    if "name" not in detected and len(all_rows) > 1:
        header_cells = all_rows[1]
        detected = _detect_columns(header_cells)
        data_rows_raw = all_rows[2:]
    else:
        data_rows_raw = all_rows[1:]

    if "name" not in detected:
        detected = {"name": 0}
        if len(header_cells) > 1:
            detected["price"] = 1
        data_rows_raw = all_rows

    return _parse_data_rows(data_rows_raw, detected)


# ---------------------------------------------------------------------------
# Shared Row Parser
# ---------------------------------------------------------------------------


def _parse_data_rows(
    data_rows: Sequence[Sequence[object]], detected: dict[str, int]
) -> ParseFileResult:
    """Parse data rows using detected column mapping."""
    rows: list[ImportRowData] = []
    skipped = 0

    name_idx = detected.get("name", 0)
    price_idx = detected.get("price")
    unit_idx = detected.get("unit")
    pack_idx = detected.get("pack_size")
    qty_idx = detected.get("qty")

    for i, row in enumerate(data_rows):
        row_no = i + 1
        cells = list(row)

        # Extract name
        if name_idx >= len(cells) or not cells[name_idx]:
            skipped += 1
            continue

        raw_name = str(cells[name_idx]).strip()
        if not raw_name or raw_name.lower() in ("none", "nan", ""):
            skipped += 1
            continue

        # Extract price
        raw_price: Decimal | None = None
        if price_idx is not None and price_idx < len(cells):
            raw_price = _parse_decimal(cells[price_idx])

        # Extract unit
        raw_unit: str | None = None
        if unit_idx is not None and unit_idx < len(cells) and cells[unit_idx]:
            raw_unit = str(cells[unit_idx]).strip()

        # Extract pack size
        raw_pack_size: Decimal | None = None
        if pack_idx is not None and pack_idx < len(cells):
            raw_pack_size = _parse_decimal(cells[pack_idx])

        # Extract qty
        raw_qty: Decimal | None = None
        if qty_idx is not None and qty_idx < len(cells):
            raw_qty = _parse_decimal(cells[qty_idx])

        # Collect extra columns
        extra: dict[str, str] = {}
        used_indices = {v for v in detected.values()}
        for col_i, cell_val in enumerate(cells):
            if col_i not in used_indices and cell_val is not None:
                extra[f"col_{col_i}"] = str(cell_val)

        rows.append(
            ImportRowData(
                row_no=row_no,
                raw_name=raw_name,
                raw_unit=raw_unit,
                raw_price=raw_price,
                raw_pack_size=raw_pack_size,
                raw_qty=raw_qty,
                extra_columns=extra,
            )
        )

    return ParseFileResult(
        rows=rows,
        total_rows=len(data_rows),
        skipped_rows=skipped,
        detected_columns=detected,
    )
