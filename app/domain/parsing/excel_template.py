"""The price-list template a shop owner downloads before uploading.

Pure: builds bytes from constants, touches no database, network or clock. The
headers here are not free text -- each one has to match a pattern in
`excel_parser._COLUMN_PATTERNS`, or the importer will ignore the column in the
very file we handed out. `test_excel_template.py` feeds the template back
through the parser to keep the two from drifting apart.

openpyxl is not fast; callers hand this to `asyncio.to_thread` rather than
building it inside a handler.
"""

from __future__ import annotations

from collections.abc import Mapping
from io import BytesIO
from typing import Final, TypeVar

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

# Column headers per language. Each must match `_COLUMN_PATTERNS` in
# excel_parser.py -- name, price, unit, pack_size, qty, in that order.
_HEADERS: Final[dict[str, tuple[str, ...]]] = {
    "uz_latn": ("Mahsulot nomi", "Birlik", "Narx (so'm)", "Qadoq", "Miqdor"),
    "uz_cyrl": ("Маҳсулот номи", "Бирлик", "Нарх (сўм)", "Қадоқ", "Миқдор"),
    "ru": ("Наименование", "Единица", "Цена (сум)", "Фасовка", "Количество"),
}

# Filled-in examples rather than empty placeholders: a shop owner copies the
# shape of a row far more reliably than they read a description of it. The
# names are real catalogue products, so an owner who edits only the price ends
# up with rows that match.
TEMPLATE_EXAMPLE_ROWS: Final[tuple[tuple[str, str, str, str, str], ...]] = (
    ("Fanera berezovaya 3x3 12 mm (1525x1525)", "dona", "157000", "1", "40"),
    ("OSB-3 plita 9 mm (2500x1250)", "dona", "118000", "1", "60"),
    ("DVP plita (T markasi) 3.2 mm (2745x1700)", "dona", "65000", "1", "25"),
)

# Notes, rendered on their own sheet. They cannot sit under the table: the
# parser skips a blank row and keeps reading, so a footer would be imported as
# three more products with no price.
_NOTES: Final[dict[str, tuple[str, ...]]] = {
    "uz_latn": (
        "Ustun nomlarini o'zgartirmang — import shu nomlar bo'yicha ishlaydi.",
        "Namunadagi 3 qatorni o'z mahsulotlaringiz bilan almashtiring.",
        "Narxni faqat raqam bilan yozing: 157000 (probel va 'so'm' shart emas).",
        "Miqdor ustuni bo'sh bo'lsa, mahsulot mavjud deb hisoblanadi.",
    ),
    "uz_cyrl": (
        "Устун номларини ўзгартирманг — импорт шу номлар бўйича ишлайди.",
        "Намунадаги 3 қаторни ўз маҳсулотларингиз билан алмаштиринг.",
        "Нархни фақат рақам билан ёзинг: 157000 (пробел ва 'сўм' шарт эмас).",
        "Миқдор устуни бўш бўлса, маҳсулот мавжуд деб ҳисобланади.",
    ),
    "ru": (
        "Не меняйте названия столбцов — импорт ориентируется на них.",
        "Замените 3 строки примера своими товарами.",
        "Цену пишите только цифрами: 157000 (без пробелов и слова «сум»).",
        "Если столбец «Количество» пуст, товар считается в наличии.",
    ),
}

_SHEET_TITLE: Final[dict[str, str]] = {
    "uz_latn": "Narxlar",
    "uz_cyrl": "Нархлар",
    "ru": "Цены",
}

_NOTES_SHEET_TITLE: Final[dict[str, str]] = {
    "uz_latn": "Yo'riqnoma",
    "uz_cyrl": "Йўриқнома",
    "ru": "Инструкция",
}

# Named so it is recognisable in a chat full of forwarded files.
TEMPLATE_FILENAME: Final[str] = "qurbot_narxlar_shabloni.xlsx"

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_COLUMN_WIDTHS: Final[tuple[int, ...]] = (46, 12, 16, 10, 12)
_DEFAULT_LANG: Final[str] = "uz_latn"


_T = TypeVar("_T")


def _for_lang(table: Mapping[str, _T], lang: str) -> _T:
    """Pick a language's entry, falling back to the default rather than raising."""
    return table.get(lang, table[_DEFAULT_LANG])


def build_price_template(lang: str = _DEFAULT_LANG) -> bytes:
    """Return a ready-to-fill price list as .xlsx bytes."""
    headers = _for_lang(_HEADERS, lang)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _for_lang(_SHEET_TITLE, lang)

    sheet.append(list(headers))
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in TEMPLATE_EXAMPLE_ROWS:
        name, unit, price, pack_size, qty = row
        sheet.append([name, unit, int(price), int(pack_size), int(qty)])

    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"

    notes_sheet = workbook.create_sheet(_for_lang(_NOTES_SHEET_TITLE, lang))
    notes_sheet.column_dimensions["A"].width = 80
    for offset, note in enumerate(_for_lang(_NOTES, lang), start=1):
        cell = notes_sheet.cell(row=offset, column=1, value=f"• {note}")
        cell.font = Font(italic=True, color="666666")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
